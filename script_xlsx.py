import os.path

from openpyxl import load_workbook

from script_os import TMP_DIR

workbook = load_workbook(os.path.join(TMP_DIR, "file_example_XLSX_50.xlsx"))
sheet = workbook.active
print(sheet.cell(row=1, column=2).value)